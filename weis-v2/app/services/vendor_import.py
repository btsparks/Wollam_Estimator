"""Vendor directory import from Master Sub List Excel format."""

import logging
from pathlib import Path
from app.database import get_connection

logger = logging.getLogger(__name__)


def _clean_phone(val) -> str | None:
    """Convert phone numbers stored as integers or strings to clean format."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == 'None':
        return None
    # Remove .0 suffix from float-stored numbers
    if s.endswith('.0'):
        s = s[:-2]
    # Remove non-digit chars for normalization, then reformat
    digits = ''.join(c for c in s if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits[0] == '1':
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return s if s else None


def _clean_str(val) -> str | None:
    """Clean a cell value to a string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != 'None' else None


def import_vendor_excel(file_path: str | Path) -> dict:
    """Parse Master Sub List Excel and upsert into vendor_directory.

    Handles 4 sheets:
    - Construction → vendor_type='construction', is_active=1
    - Materials & Suppliers → vendor_type='materials', is_active=1
    - Construction (Old) → vendor_type='construction', is_active=0
    - Materials & Suppliers (Old) → vendor_type='materials', is_active=0
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    sheet_config = [
        ("Construction", "construction", True),
        ("Materials & Suppliers", "materials", True),
        ("Construction (Old)", "construction", False),
        ("Materials & Suppliers (Old)", "materials", False),
    ]

    conn = get_connection()
    created = 0
    updated = 0
    skipped = 0

    try:
        for sheet_name, vendor_type, is_active in sheet_config:
            if sheet_name not in wb.sheetnames:
                logger.info("Sheet '%s' not found, skipping", sheet_name)
                continue

            ws = wb[sheet_name]
            current_trade = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue

                # Column A = trade (sticky — carries forward until next trade header)
                trade_val = _clean_str(row[0])
                if trade_val:
                    current_trade = trade_val

                # Column B = company (required)
                company = _clean_str(row[1]) if len(row) > 1 else None
                if not company or not current_trade:
                    continue

                contact = _clean_str(row[2]) if len(row) > 2 else None
                city = _clean_str(row[3]) if len(row) > 3 else None
                state = _clean_str(row[4]) if len(row) > 4 else None
                phone = _clean_phone(row[5]) if len(row) > 5 else None
                cell = _clean_phone(row[6]) if len(row) > 6 else None
                email = _clean_str(row[7]) if len(row) > 7 else None
                website = _clean_str(row[8]) if len(row) > 8 else None
                fax = _clean_phone(row[9]) if len(row) > 9 else None

                # Column K = DBE
                is_dbe = 0
                if len(row) > 10 and row[10]:
                    dbe_val = str(row[10]).strip().upper()
                    is_dbe = 1 if dbe_val in ('Y', 'YES', '1', 'TRUE', 'X') else 0

                specialties = _clean_str(row[11]) if len(row) > 11 else None

                # Materials sheets may have second contact in cols L-N
                second_contact = None
                second_phone = None
                second_email = None
                if vendor_type == 'materials' and len(row) > 13:
                    second_contact = _clean_str(row[11])
                    second_phone = _clean_phone(row[12])
                    second_email = _clean_str(row[13])
                    specialties = None  # Materials uses those cols for 2nd contact

                # Upsert: dedup on (company, trade)
                existing = conn.execute(
                    "SELECT id FROM vendor_directory WHERE company = ? AND trade = ?",
                    (company, current_trade),
                ).fetchone()

                if existing:
                    conn.execute(
                        """UPDATE vendor_directory SET vendor_type=?, contact_name=?, city=?, state=?,
                           phone=?, cell=?, email=?, website=?, fax=?, is_dbe=?, specialties=?,
                           second_contact=?, second_phone=?, second_email=?, is_active=?,
                           updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                        (vendor_type, contact, city, state, phone, cell, email, website, fax,
                         is_dbe, specialties, second_contact, second_phone, second_email,
                         1 if is_active else 0, existing["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        """INSERT INTO vendor_directory
                           (vendor_type, trade, company, contact_name, city, state, phone, cell,
                            email, website, fax, is_dbe, specialties, second_contact, second_phone,
                            second_email, is_active)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (vendor_type, current_trade, company, contact, city, state, phone, cell,
                         email, website, fax, is_dbe, specialties, second_contact, second_phone,
                         second_email, 1 if is_active else 0),
                    )
                    created += 1

        conn.commit()
        wb.close()
        return {"status": "ok", "created": created, "updated": updated, "skipped": skipped}
    except Exception:
        conn.rollback()
        wb.close()
        raise
    finally:
        conn.close()
