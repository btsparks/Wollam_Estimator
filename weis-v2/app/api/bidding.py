"""Bidding Platform API — Bid Board, Documents, Schedule of Values, Pricing Groups.

Separate from estimates.py which serves HeavyBid historicals.
"""

from __future__ import annotations

import hashlib
import logging
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.config import BID_DOCUMENTS_DIR, ESTIMATING_ROOT, DROPBOX_ROOT, ANTHROPIC_API_KEY, VECTOR_SEARCH_ENABLED
from app.database import get_connection
from app.services.document_extract import extract_text
from app.services.sov_parser import parse_sov_file
from app.services.bid_sync import resolve_bid_folder, sync_bid_documents, discover_bid_files
from app.services.document_chunker import chunk_document, chunk_all_bid_documents
from app.services.rate_lookup import lookup_rates_for_sov_item, auto_populate_sov_rates

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/bidding", tags=["bidding"])

ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv", ".txt", ".docx", ".doc"}

DOC_CATEGORIES = {
    "spec", "drawing", "contract", "bid_schedule", "rfi_clarification",
    "addendum_package", "bond_form", "insurance", "general",
}


# ── Request/Response Models ─────────────────────────────────────

class BidCreate(BaseModel):
    bid_name: str
    bid_number: Optional[str] = None
    owner: Optional[str] = None
    general_contractor: Optional[str] = None
    bid_date: Optional[str] = None
    bid_due_time: Optional[str] = None
    project_type: Optional[str] = None
    location: Optional[str] = None
    estimated_value: Optional[float] = None
    description: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    status: Optional[str] = "active"
    notes: Optional[str] = None
    dropbox_folder_path: Optional[str] = None


class BidUpdate(BaseModel):
    bid_name: Optional[str] = None
    bid_number: Optional[str] = None
    owner: Optional[str] = None
    general_contractor: Optional[str] = None
    bid_date: Optional[str] = None
    bid_due_time: Optional[str] = None
    project_type: Optional[str] = None
    location: Optional[str] = None
    estimated_value: Optional[float] = None
    description: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None


class DocUpdate(BaseModel):
    doc_category: Optional[str] = None
    doc_label: Optional[str] = None
    addendum_number: Optional[int] = None
    date_received: Optional[str] = None
    notes: Optional[str] = None


class SOVItemCreate(BaseModel):
    item_number: Optional[str] = None
    description: str
    quantity: Optional[float] = None
    unit: Optional[str] = None
    notes: Optional[str] = None
    pricing_group_id: Optional[int] = None
    hcss_number: Optional[str] = None
    work_type: Optional[str] = None
    section_id: Optional[int] = None
    is_holding_account: Optional[int] = None
    holding_description: Optional[str] = None


class SOVItemUpdate(BaseModel):
    item_number: Optional[str] = None
    description: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    notes: Optional[str] = None
    pricing_group_id: Optional[int] = None
    in_scope: Optional[int] = None
    hcss_number: Optional[str] = None
    work_type: Optional[str] = None
    section_id: Optional[int] = None
    is_holding_account: Optional[int] = None
    holding_description: Optional[str] = None


class SOVScopeUpdate(BaseModel):
    item_ids: list[int]
    in_scope: int  # 1 or 0


class SOVWorkTypeUpdate(BaseModel):
    item_ids: list[int]
    work_type: str  # 'self_perform', 'subcontract', or 'undecided'


class SOVConfirm(BaseModel):
    items: list[dict]


class ReorderRequest(BaseModel):
    item_ids: list[int]


class GroupCreate(BaseModel):
    name: str
    description: Optional[str] = None


class GroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


class AssignItems(BaseModel):
    item_ids: list[int]


class SectionCreate(BaseModel):
    name: str
    sort_order: Optional[float] = None


class SectionUpdate(BaseModel):
    name: Optional[str] = None
    sort_order: Optional[float] = None
    collapsed: Optional[int] = None


class SectionAssign(BaseModel):
    item_ids: list[int]


class HoldingAccountCreate(BaseModel):
    holding_description: Optional[str] = None


class HoldingDistribute(BaseModel):
    target_item_ids: list[int]


class DrawingRegisterUpdate(BaseModel):
    drawing_number: Optional[str] = None
    title: Optional[str] = None
    discipline: Optional[str] = None
    revision: Optional[str] = None
    notes: Optional[str] = None


class SpecRegisterUpdate(BaseModel):
    spec_section: Optional[str] = None
    title: Optional[str] = None
    division: Optional[str] = None
    notes: Optional[str] = None


class RFICreate(BaseModel):
    rfi_number: Optional[str] = None
    question: str
    asked_by: Optional[str] = None
    date_asked: Optional[str] = None
    response: Optional[str] = None
    responded_by: Optional[str] = None
    date_responded: Optional[str] = None
    addendum_number: Optional[int] = None
    related_spec: Optional[str] = None
    related_drawing: Optional[str] = None
    status: Optional[str] = "answered"
    notes: Optional[str] = None


class RFIUpdate(BaseModel):
    rfi_number: Optional[str] = None
    question: Optional[str] = None
    asked_by: Optional[str] = None
    date_asked: Optional[str] = None
    response: Optional[str] = None
    responded_by: Optional[str] = None
    date_responded: Optional[str] = None
    addendum_number: Optional[int] = None
    related_spec: Optional[str] = None
    related_drawing: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None


# ── Bid CRUD ─────────────────────────────────────────────────────

@router.get("/bids")
async def list_bids(status: Optional[str] = Query(None)):
    """List all bids with optional status filter."""
    conn = get_connection()
    try:
        if status:
            bids = conn.execute(
                "SELECT * FROM active_bids WHERE status = ? ORDER BY bid_date ASC",
                (status,),
            ).fetchall()
        else:
            bids = conn.execute(
                "SELECT * FROM active_bids ORDER BY bid_date ASC"
            ).fetchall()

        result = []
        for b in bids:
            bid = dict(b)
            # Add doc count and sov count
            doc_count = conn.execute(
                "SELECT COUNT(*) as cnt FROM bid_documents WHERE bid_id = ?",
                (b["id"],),
            ).fetchone()["cnt"]
            sov_count = conn.execute(
                "SELECT COUNT(*) as cnt FROM bid_sov_item WHERE bid_id = ?",
                (b["id"],),
            ).fetchone()["cnt"]
            bid["doc_count"] = doc_count
            bid["sov_count"] = sov_count
            result.append(bid)

        return result
    finally:
        conn.close()


@router.post("/bids")
async def create_bid(bid: BidCreate):
    """Create a new bid project."""
    conn = get_connection()
    try:
        # Auto-generate bid number if not provided: YY-MM-NNNN
        bid_number = bid.bid_number
        if not bid_number:
            now = datetime.now()
            prefix = now.strftime("%y-%m")
            # Find the highest sequential number across all bids
            row = conn.execute(
                """SELECT bid_number FROM active_bids
                   WHERE bid_number LIKE '%-%-____'
                   ORDER BY CAST(SUBSTR(bid_number, 7) AS INTEGER) DESC
                   LIMIT 1"""
            ).fetchone()
            if row and row["bid_number"]:
                try:
                    last_seq = int(row["bid_number"].split("-")[-1])
                    bid_number = f"{prefix}-{last_seq + 1}"
                except (ValueError, IndexError):
                    bid_number = f"{prefix}-0001"
            else:
                bid_number = f"{prefix}-0001"

        cursor = conn.execute(
            """INSERT INTO active_bids
               (bid_name, bid_number, owner, general_contractor, bid_date, bid_due_time,
                project_type, location, estimated_value, description, contact_name,
                contact_email, status, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                bid.bid_name, bid_number, bid.owner, bid.general_contractor,
                bid.bid_date, bid.bid_due_time, bid.project_type, bid.location,
                bid.estimated_value, bid.description, bid.contact_name,
                bid.contact_email, bid.status, bid.notes,
            ),
        )
        conn.commit()
        bid_id = cursor.lastrowid

        # Link Dropbox folder if path provided by user
        if bid.dropbox_folder_path:
            clean_path = bid.dropbox_folder_path.strip().strip('"').strip("'")
            conn.execute(
                "UPDATE active_bids SET dropbox_folder_path = ? WHERE id = ?",
                (clean_path, bid_id),
            )
            conn.commit()

        row = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.get("/bids/{bid_id}")
async def get_bid(bid_id: int):
    """Get bid detail with counts."""
    conn = get_connection()
    try:
        row = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Bid not found")

        bid = dict(row)
        bid["doc_count"] = conn.execute(
            "SELECT COUNT(*) as cnt FROM bid_documents WHERE bid_id = ?", (bid_id,)
        ).fetchone()["cnt"]
        bid["sov_count"] = conn.execute(
            "SELECT COUNT(*) as cnt FROM bid_sov_item WHERE bid_id = ?", (bid_id,)
        ).fetchone()["cnt"]
        bid["group_count"] = conn.execute(
            "SELECT COUNT(*) as cnt FROM pricing_group WHERE bid_id = ?", (bid_id,)
        ).fetchone()["cnt"]
        return bid
    finally:
        conn.close()


@router.put("/bids/{bid_id}")
async def update_bid(bid_id: int, bid: BidUpdate):
    """Update bid fields."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Bid not found")

        updates = {k: v for k, v in bid.model_dump().items() if v is not None}
        if not updates:
            return dict(existing)

        updates["updated_at"] = datetime.now(tz=timezone.utc).isoformat()
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [bid_id]

        conn.execute(f"UPDATE active_bids SET {set_clause} WHERE id = ?", values)
        conn.commit()

        row = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/bids/{bid_id}")
async def delete_bid(bid_id: int):
    """Delete bid and all children (cascade)."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Bid not found")

        # Clean up vector embeddings
        if VECTOR_SEARCH_ENABLED:
            try:
                from app.services.vector_store import delete_bid_collection
                delete_bid_collection(bid_id)
            except Exception as e:
                logger.warning("Failed to delete vector collection for bid %d: %s", bid_id, e)

        # Delete children first (foreign key order)
        conn.execute("DELETE FROM rfi_log WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM drawing_register WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM spec_register WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM bid_document_chunks WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM bid_documents WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM holding_distribution WHERE holding_item_id IN (SELECT id FROM bid_sov_item WHERE bid_id = ?)", (bid_id,))
        conn.execute("DELETE FROM bid_sov_item WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM bid_section WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM pricing_group WHERE bid_id = ?", (bid_id,))
        conn.execute("DELETE FROM active_bids WHERE id = ?", (bid_id,))
        conn.commit()

        # Clean up files
        bid_dir = BID_DOCUMENTS_DIR / str(bid_id)
        if bid_dir.exists():
            shutil.rmtree(bid_dir)

        return {"deleted": True, "bid_id": bid_id}
    finally:
        conn.close()


# ── Document Management ──────────────────────────────────────────

@router.post("/bids/{bid_id}/documents")
async def upload_document(
    bid_id: int,
    file: UploadFile = File(...),
    addendum_number: int = Form(0),
    doc_category: str = Form("general"),
    date_received: Optional[str] = Form(None),
):
    """Upload a document with metadata."""
    conn = get_connection()
    try:
        # Verify bid exists
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        # Validate file type
        suffix = Path(file.filename).suffix.lower()
        if suffix not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {suffix}. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
            )

        # Validate category
        if doc_category not in DOC_CATEGORIES:
            doc_category = "general"

        # Save file to disk
        bid_dir = BID_DOCUMENTS_DIR / str(bid_id)
        bid_dir.mkdir(parents=True, exist_ok=True)

        file_path = bid_dir / file.filename
        content = await file.read()
        file_path.write_bytes(content)

        file_hash = hashlib.sha256(content).hexdigest()
        file_size = len(content)

        # Extract text
        extracted_text = ""
        extraction_status = "pending"
        extraction_warning = None
        page_count = None
        word_count = None

        try:
            extracted_text = extract_text(file_path)
            extraction_status = "complete"
            word_count = len(extracted_text.split()) if extracted_text else 0
            if suffix == ".pdf":
                page_count = extracted_text.count("--- Page ")
        except Exception as e:
            extraction_status = "error"
            extraction_warning = str(e)
            logger.warning(f"Text extraction failed for {file.filename}: {e}")

        # Insert document record
        cursor = conn.execute(
            """INSERT INTO bid_documents
               (bid_id, filename, file_type, file_size_bytes, doc_category,
                extraction_status, extraction_warning, page_count, word_count,
                file_hash, addendum_number, date_received, file_path, extracted_text)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                bid_id, file.filename, suffix, file_size, doc_category,
                extraction_status, extraction_warning, page_count, word_count,
                file_hash, addendum_number, date_received, str(file_path),
                extracted_text,
            ),
        )
        conn.commit()

        doc_id = cursor.lastrowid
        conn.close()
        conn = None

        # Chunk the document for agent analysis
        if extraction_status == "complete":
            try:
                chunk_document(doc_id)
            except Exception as e:
                logger.warning("Chunking failed for doc %d: %s", doc_id, e)

            # Embed chunks into vector store
            if VECTOR_SEARCH_ENABLED:
                try:
                    from app.services.vector_store import embed_document_chunks
                    embed_document_chunks(bid_id, doc_id)
                except Exception as e:
                    logger.warning("Embedding failed for doc %d: %s", doc_id, e)

        conn = get_connection()
        row = conn.execute("SELECT * FROM bid_documents WHERE id = ?", (doc_id,)).fetchone()
        result = dict(row)
        # Don't send full extracted text in response
        result.pop("extracted_text", None)
        return result
    finally:
        if conn:
            conn.close()


@router.get("/bids/{bid_id}/documents")
async def list_documents(
    bid_id: int,
    addendum_number: Optional[int] = Query(None),
    doc_category: Optional[str] = Query(None),
    sync_action: Optional[str] = Query(None),
):
    """List documents for a bid, filterable by addendum, category, or sync action."""
    conn = get_connection()
    try:
        query = "SELECT * FROM bid_documents WHERE bid_id = ?"
        params: list = [bid_id]

        if addendum_number is not None:
            query += " AND addendum_number = ?"
            params.append(addendum_number)
        if doc_category:
            query += " AND doc_category = ?"
            params.append(doc_category)
        if sync_action:
            query += " AND sync_action = ?"
            params.append(sync_action)

        query += " ORDER BY addendum_number ASC, created_at ASC"

        rows = conn.execute(query, params).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            # Don't send full extracted text in list view
            d.pop("extracted_text", None)
            result.append(d)
        return result
    finally:
        conn.close()


@router.get("/bids/{bid_id}/documents/tree")
async def get_document_tree(bid_id: int):
    """Get documents organized in a folder tree structure."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM bid_documents WHERE bid_id = ? ORDER BY addendum_number ASC, created_at ASC",
            (bid_id,),
        ).fetchall()

        # Build tree from dropbox_path segments or doc_category
        tree: dict = {}  # path -> {type, name, path, children/doc}

        for r in rows:
            doc = dict(r)
            doc.pop("extracted_text", None)
            doc.pop("previous_extracted_text", None)

            # Build file:// URL from file_path or dropbox_path
            file_path = doc.get("file_path") or doc.get("dropbox_path") or ""
            if file_path:
                doc["dropbox_url"] = "file:///" + file_path.replace("\\", "/").replace(" ", "%20")
            else:
                doc["dropbox_url"] = None

            doc["is_new"] = doc.get("sync_action") == "new"

            # Determine folder path
            dropbox_path = doc.get("dropbox_path") or ""
            if dropbox_path:
                # Parse path segments relative to bid folder
                # dropbox_path is absolute — extract relative portion after bid folder
                parts = dropbox_path.replace("\\", "/").split("/")
                # Find the filename and use parent segments as folder path
                filename = parts[-1] if parts else doc["filename"]
                folder_parts = parts[:-1] if len(parts) > 1 else []

                # Use only the last 2-3 path segments (within the bid folder)
                # Look for the bid folder name pattern (YY-MM-NNNN)
                bid_folder_idx = -1
                for i, p in enumerate(folder_parts):
                    import re
                    if re.match(r"\d{2}-\d{2}-\d{4}", p):
                        bid_folder_idx = i
                        break

                if bid_folder_idx >= 0:
                    folder_parts = folder_parts[bid_folder_idx + 1:]
                else:
                    # Use last 2 parts as relative path
                    folder_parts = folder_parts[-2:] if len(folder_parts) > 2 else folder_parts

                folder_key = "/".join(folder_parts) if folder_parts else ""
            else:
                # Use doc_category as folder for manually uploaded docs
                cat = doc.get("doc_category") or "general"
                folder_key = cat.replace("_", " ").title()

            # Insert into tree structure
            if folder_key not in tree:
                tree[folder_key] = {"type": "folder", "name": folder_key.split("/")[-1] if folder_key else "Root", "path": folder_key, "children": [], "doc_count": 0, "has_new": False, "has_updated": False}
            tree[folder_key]["children"].append({
                "type": "document",
                "id": doc["id"],
                "filename": doc["filename"],
                "doc_category": doc.get("doc_category"),
                "addendum_number": doc.get("addendum_number", 0),
                "date_received": doc.get("date_received"),
                "sync_action": doc.get("sync_action"),
                "is_new": doc.get("is_new", False),
                "word_count": doc.get("word_count", 0),
                "file_size_bytes": doc.get("file_size_bytes", 0),
                "dropbox_url": doc.get("dropbox_url"),
                "file_path": file_path,
                "version": doc.get("version", 1),
            })
            tree[folder_key]["doc_count"] += 1
            if doc.get("sync_action") == "new":
                tree[folder_key]["has_new"] = True
            if doc.get("sync_action") == "updated":
                tree[folder_key]["has_updated"] = True

        # Build nested tree from flat folder paths
        def build_nested(flat_tree):
            nested = {}
            for path, folder in sorted(flat_tree.items()):
                parts = path.split("/") if path else [""]
                current = nested
                for i, part in enumerate(parts):
                    subpath = "/".join(parts[:i+1])
                    if subpath not in current:
                        current[subpath] = {"type": "folder", "name": part or "Root", "path": subpath, "children": [], "subfolders": {}, "doc_count": 0, "has_new": False, "has_updated": False}
                    if i == len(parts) - 1:
                        current[subpath]["children"] = folder["children"]
                        current[subpath]["doc_count"] = folder["doc_count"]
                        current[subpath]["has_new"] = folder["has_new"]
                        current[subpath]["has_updated"] = folder["has_updated"]
                    current = current[subpath].get("subfolders", {})
                    if not isinstance(current, dict):
                        current = {}
            return nested

        # Flatten to sorted list
        sorted_folders = sorted(tree.values(), key=lambda f: f["path"])

        # Stats
        total = len(rows)
        new_count = sum(1 for r in rows if dict(r).get("sync_action") == "new")
        updated_count = sum(1 for r in rows if dict(r).get("sync_action") == "updated")

        return {
            "tree": sorted_folders,
            "stats": {
                "total_documents": total,
                "new_since_last_sync": new_count,
                "updated_since_last_sync": updated_count,
            },
        }
    finally:
        conn.close()


@router.get("/documents/{doc_id}")
async def get_document(doc_id: int):
    """Get single document detail including extracted text."""
    conn = get_connection()
    try:
        row = conn.execute("SELECT * FROM bid_documents WHERE id = ?", (doc_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Document not found")
        return dict(row)
    finally:
        conn.close()


@router.put("/documents/{doc_id}")
async def update_document(doc_id: int, update: DocUpdate):
    """Update document metadata."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_documents WHERE id = ?", (doc_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Document not found")

        updates = {k: v for k, v in update.model_dump().items() if v is not None}
        if not updates:
            result = dict(existing)
            result.pop("extracted_text", None)
            return result

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [doc_id]
        conn.execute(f"UPDATE bid_documents SET {set_clause} WHERE id = ?", values)
        conn.commit()

        row = conn.execute("SELECT * FROM bid_documents WHERE id = ?", (doc_id,)).fetchone()
        result = dict(row)
        result.pop("extracted_text", None)
        return result
    finally:
        conn.close()


@router.delete("/documents/{doc_id}")
async def delete_document(doc_id: int):
    """Delete document and its chunks."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_documents WHERE id = ?", (doc_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Document not found")

        # Remove vector embeddings
        if VECTOR_SEARCH_ENABLED:
            try:
                from app.services.vector_store import remove_document_embeddings
                remove_document_embeddings(existing["bid_id"], doc_id)
            except Exception as e:
                logger.warning("Failed to remove embeddings for doc %d: %s", doc_id, e)

        conn.execute("DELETE FROM bid_document_chunks WHERE document_id = ?", (doc_id,))
        conn.execute("DELETE FROM bid_documents WHERE id = ?", (doc_id,))
        conn.commit()

        # Clean up file — only delete locally-uploaded files, never Dropbox source files
        if existing["file_path"] and not existing["dropbox_path"]:
            fp = Path(existing["file_path"])
            if fp.exists():
                fp.unlink()

        return {"deleted": True, "doc_id": doc_id}
    finally:
        conn.close()


# ── Dropbox Folder Selection ───────────────────────────────────

@router.post("/pick-folder")
async def pick_folder():
    """Open a native OS folder picker dialog and return the selected path.

    Runs tkinter in a separate subprocess so it can never block the server.
    """
    import subprocess
    import sys

    initial_dir = str(ESTIMATING_ROOT) if ESTIMATING_ROOT.exists() else str(DROPBOX_ROOT)

    script = f"""
import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()
root.attributes("-topmost", True)
root.focus_force()
folder = filedialog.askdirectory(title="Select Estimating Folder", initialdir=r"{initial_dir}")
root.destroy()
print(folder if folder else "")
"""

    try:
        proc = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, timeout=120,
        )
        path = proc.stdout.strip()
        if not path:
            return {"picked": False, "message": "No folder selected"}

        folder = Path(path)
        return {"picked": True, "path": str(folder), "name": folder.name}
    except subprocess.TimeoutExpired:
        return {"picked": False, "message": "Folder picker timed out"}
    except Exception as e:
        return {"picked": False, "message": str(e)}


@router.get("/browse-folders")
async def browse_folders(q: str = Query(default="")):
    """List available Dropbox estimating folders for manual selection.

    Returns all subfolders in ESTIMATING_ROOT, optionally filtered by query string.
    """
    if not ESTIMATING_ROOT.exists():
        return {"folders": [], "root": str(ESTIMATING_ROOT), "error": "Estimating root not found"}

    folders = []
    for entry in sorted(ESTIMATING_ROOT.iterdir()):
        if entry.is_dir() and not entry.name.startswith("."):
            folders.append({"name": entry.name, "path": str(entry)})

    # Filter by query if provided
    if q.strip():
        query = q.strip().lower()
        folders = [f for f in folders if query in f["name"].lower()]

    return {"folders": folders, "root": str(ESTIMATING_ROOT)}


class ResolveFolderName(BaseModel):
    folder_name: str


@router.post("/resolve-folder-name")
async def resolve_folder_name(body: ResolveFolderName):
    """Resolve a folder name (from native OS picker) to its full path.

    Searches in order: ESTIMATING_ROOT, then DROPBOX_ROOT.
    """
    name = body.folder_name.strip()
    if not name:
        return {"found": False, "message": "No folder name provided"}

    # Search ESTIMATING_ROOT first (Estimates - Shared), then DROPBOX_ROOT
    search_roots = []
    if ESTIMATING_ROOT.exists():
        search_roots.append(ESTIMATING_ROOT)
    if DROPBOX_ROOT.exists() and DROPBOX_ROOT != ESTIMATING_ROOT:
        search_roots.append(DROPBOX_ROOT)

    for root in search_roots:
        # Exact match
        target = root / name
        if target.exists() and target.is_dir():
            return {"found": True, "path": str(target), "name": name}

        # Case-insensitive match
        try:
            for entry in root.iterdir():
                if entry.is_dir() and entry.name.lower() == name.lower():
                    return {"found": True, "path": str(entry), "name": entry.name}
        except PermissionError:
            continue

    return {"found": False, "message": f"Folder '{name}' not found in Dropbox. You can paste the full folder path instead."}


class LinkFolderRequest(BaseModel):
    folder_path: Optional[str] = None


@router.post("/bids/{bid_id}/link-folder")
async def link_folder(bid_id: int, body: LinkFolderRequest):
    """Link a Dropbox estimating folder to a bid by explicit path."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        if not body.folder_path:
            return {"linked": False, "message": "No folder path provided"}

        folder = Path(body.folder_path.strip().strip('"').strip("'"))
        if not folder.exists() or not folder.is_dir():
            return {"linked": False, "message": f"Folder not found: {body.folder_path}"}

        conn.execute(
            "UPDATE active_bids SET dropbox_folder_path = ? WHERE id = ?",
            (str(folder), bid_id),
        )
        conn.commit()
        return {"linked": True, "folder_path": str(folder)}
    finally:
        conn.close()


@router.post("/bids/{bid_id}/sync")
async def sync_documents(bid_id: int):
    """Trigger a Dropbox folder sync for a bid."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id, dropbox_folder_path FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")
        if not bid["dropbox_folder_path"]:
            raise HTTPException(status_code=400, detail="No Dropbox folder linked to this bid")
    finally:
        conn.close()

    try:
        result = sync_bid_documents(bid_id)
        return result
    except Exception as e:
        logger.error("Sync failed for bid %d: %s", bid_id, e)
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


@router.get("/bids/{bid_id}/sync-stream")
async def sync_documents_stream(bid_id: int):
    """Stream sync progress via Server-Sent Events.

    Each file processed sends an SSE event with current/total/filename/action.
    Final event has type 'done' with the full counts.
    """
    import asyncio
    import json as _json
    import queue
    import threading

    conn = get_connection()
    try:
        bid = conn.execute("SELECT id, dropbox_folder_path FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")
        if not bid["dropbox_folder_path"]:
            raise HTTPException(status_code=400, detail="No Dropbox folder linked to this bid")
    finally:
        conn.close()

    progress_queue = queue.Queue()

    def on_progress(current, total, filename, action):
        progress_queue.put({"type": "progress", "current": current, "total": total, "filename": filename, "action": action})

    def run_sync():
        try:
            result = sync_bid_documents(bid_id, on_progress=on_progress)
            progress_queue.put({"type": "done", "result": result})
        except Exception as e:
            progress_queue.put({"type": "error", "message": str(e)})

    thread = threading.Thread(target=run_sync, daemon=True)
    thread.start()

    async def event_stream():
        while True:
            try:
                msg = progress_queue.get_nowait()
                yield f"data: {_json.dumps(msg)}\n\n"
                if msg["type"] in ("done", "error"):
                    return
            except queue.Empty:
                await asyncio.sleep(0.3)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/bids/{bid_id}/sync-status")
async def get_sync_status(bid_id: int):
    """Get sync status and document counts for a bid."""
    conn = get_connection()
    try:
        bid = conn.execute(
            "SELECT sync_status, last_synced_at FROM active_bids WHERE id = ?",
            (bid_id,),
        ).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        # Count documents by sync_action
        action_counts = conn.execute(
            """SELECT sync_action, COUNT(*) as cnt
               FROM bid_documents
               WHERE bid_id = ? AND sync_action IS NOT NULL
               GROUP BY sync_action""",
            (bid_id,),
        ).fetchall()

        document_counts = {r["sync_action"]: r["cnt"] for r in action_counts}
        document_counts["total"] = sum(document_counts.values())

        return {
            "sync_status": bid["sync_status"],
            "last_synced_at": bid["last_synced_at"],
            "document_counts": document_counts,
        }
    finally:
        conn.close()


# ── AI Overview Analysis ────────────────────────────────────────

@router.post("/bids/{bid_id}/analyze-overview")
async def analyze_overview(bid_id: int):
    """AI analyzes synced documents to auto-populate bid overview fields.

    Reads extracted text from the bid's documents, sends to Claude to extract
    project metadata (owner, GC, location, type, description), and updates
    the bid record with any fields not already filled in by the user.
    """
    import anthropic
    import json as _json

    conn = get_connection()
    try:
        bid = conn.execute("SELECT * FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        # Get documents — prioritize contracts, specs, bid schedules
        docs = conn.execute(
            """SELECT doc_label, doc_category, extracted_text
               FROM bid_documents
               WHERE bid_id = ? AND extracted_text IS NOT NULL AND extracted_text != ''
               ORDER BY CASE doc_category
                   WHEN 'contract' THEN 1
                   WHEN 'bid_schedule' THEN 2
                   WHEN 'spec' THEN 3
                   WHEN 'addendum_package' THEN 4
                   WHEN 'general' THEN 5
                   ELSE 6
               END
               LIMIT 10""",
            (bid_id,),
        ).fetchall()

        if not docs:
            return {"analyzed": False, "message": "No documents with extracted text found. Sync documents first."}

        # Build context from documents — cap total at 60K chars
        doc_texts = []
        total_chars = 0
        for doc in docs:
            text = doc["extracted_text"]
            remaining = 60000 - total_chars
            if remaining <= 0:
                break
            if len(text) > remaining:
                text = text[:remaining]
            doc_texts.append(f"--- Document: {doc['doc_label']} (Category: {doc['doc_category']}) ---\n{text}")
            total_chars += len(text)

        combined = "\n\n".join(doc_texts)

        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": f"""Analyze these construction bid documents and extract project information. Return a JSON object with ONLY these fields (use null for anything you can't determine):

{{
  "owner": "the project owner or agency (who is paying for the work)",
  "general_contractor": "the general contractor or prime contractor (if this is a sub-bid)",
  "location": "project location (city, state or address)",
  "project_type": "type of work (e.g. heavy civil, industrial, pipeline, building, earthwork)",
  "description": "2-3 sentence description of the project scope",
  "estimated_value": null,
  "bid_date": null
}}

For bid_date, use YYYY-MM-DD format if you find a bid due date. For estimated_value, use a number (no $ or commas) if you find an engineer's estimate.

Return ONLY valid JSON, no other text.

Documents:
{combined}"""
            }],
        )

        # Parse AI response
        try:
            raw = response.content[0].text.strip()
            # Strip markdown code fence if present
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
                if raw.endswith("```"):
                    raw = raw[:-3]
                raw = raw.strip()
            extracted = _json.loads(raw)
        except (_json.JSONDecodeError, IndexError):
            logger.warning("AI overview analysis returned invalid JSON for bid %d", bid_id)
            return {"analyzed": False, "message": "AI could not parse documents into structured data"}

        # Update bid — only fill fields that are currently empty
        updates = []
        params = []
        field_map = {
            "owner": "owner",
            "general_contractor": "general_contractor",
            "location": "location",
            "project_type": "project_type",
            "description": "description",
            "estimated_value": "estimated_value",
            "bid_date": "bid_date",
        }

        populated = []
        for ai_key, db_col in field_map.items():
            ai_val = extracted.get(ai_key)
            current_val = bid[db_col]
            if ai_val is not None and (current_val is None or str(current_val).strip() == ""):
                updates.append(f"{db_col} = ?")
                params.append(ai_val if not isinstance(ai_val, float) else ai_val)
                populated.append(db_col)

        if updates:
            params.append(bid_id)
            conn.execute(
                f"UPDATE active_bids SET {', '.join(updates)} WHERE id = ?",
                params,
            )
            conn.commit()

        return {
            "analyzed": True,
            "fields_populated": populated,
            "extracted": extracted,
        }
    finally:
        conn.close()


# ── Schedule of Values ───────────────────────────────────────────

@router.post("/bids/{bid_id}/sov/upload")
async def upload_sov(bid_id: int, file: UploadFile = File(...)):
    """Upload bid schedule file, AI-parse it, return preview."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")
    finally:
        conn.close()

    # Save temp file for parsing
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {suffix}")

    bid_dir = BID_DOCUMENTS_DIR / str(bid_id)
    bid_dir.mkdir(parents=True, exist_ok=True)
    temp_path = bid_dir / f"_sov_upload_{file.filename}"

    content = await file.read()
    temp_path.write_bytes(content)

    try:
        items = parse_sov_file(temp_path)
        return {"items": items, "count": len(items), "filename": file.filename}
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        # Clean up temp file
        if temp_path.exists():
            temp_path.unlink()


@router.post("/bids/{bid_id}/sov/confirm")
async def confirm_sov(bid_id: int, body: SOVConfirm):
    """Confirm parsed SOV items and save to database."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        # Get current max sort_order
        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) as m FROM bid_sov_item WHERE bid_id = ?",
            (bid_id,),
        ).fetchone()["m"]

        saved = 0
        for i, item in enumerate(body.items):
            conn.execute(
                """INSERT INTO bid_sov_item
                   (bid_id, item_number, description, quantity, unit, sort_order, mapped_by)
                   VALUES (?, ?, ?, ?, ?, ?, 'ai_parsed')""",
                (
                    bid_id,
                    item.get("item_number"),
                    item.get("description", ""),
                    item.get("quantity"),
                    item.get("unit"),
                    max_sort + 1 + i,
                ),
            )
            saved += 1

        conn.commit()
        return {"saved": saved, "bid_id": bid_id}
    finally:
        conn.close()


@router.get("/bids/{bid_id}/sov")
async def list_sov(bid_id: int):
    """Get all SOV items for a bid, grouped by section."""
    conn = get_connection()
    try:
        rows = conn.execute(
            """SELECT s.*, pg.name as group_name,
                      bs.name as section_name, bs.sort_order as section_sort_order,
                      bs.collapsed as section_collapsed
               FROM bid_sov_item s
               LEFT JOIN pricing_group pg ON s.pricing_group_id = pg.id
               LEFT JOIN bid_section bs ON s.section_id = bs.id
               WHERE s.bid_id = ?
               ORDER BY s.sort_order ASC""",
            (bid_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.post("/bids/{bid_id}/sov")
async def add_sov_item(bid_id: int, item: SOVItemCreate):
    """Manually add a SOV item."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) as m FROM bid_sov_item WHERE bid_id = ?",
            (bid_id,),
        ).fetchone()["m"]

        cursor = conn.execute(
            """INSERT INTO bid_sov_item
               (bid_id, item_number, description, quantity, unit, notes,
                pricing_group_id, hcss_number, work_type, section_id,
                is_holding_account, holding_description, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                bid_id, item.item_number, item.description, item.quantity,
                item.unit, item.notes, item.pricing_group_id,
                item.hcss_number, item.work_type or 'undecided', item.section_id,
                item.is_holding_account or 0, item.holding_description, max_sort + 1,
            ),
        )
        conn.commit()

        row = conn.execute(
            "SELECT * FROM bid_sov_item WHERE id = ?", (cursor.lastrowid,)
        ).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.put("/sov/reorder")
async def reorder_sov(body: ReorderRequest):
    """Reorder SOV items — item_ids in desired order."""
    conn = get_connection()
    try:
        for i, item_id in enumerate(body.item_ids):
            conn.execute(
                "UPDATE bid_sov_item SET sort_order = ? WHERE id = ?",
                (i, item_id),
            )
        conn.commit()
        return {"reordered": len(body.item_ids)}
    finally:
        conn.close()


@router.post("/bids/{bid_id}/sov/set-scope")
async def set_sov_scope(bid_id: int, body: SOVScopeUpdate):
    """Set in_scope flag for multiple SOV items at once."""
    conn = get_connection()
    try:
        placeholders = ",".join("?" for _ in body.item_ids)
        conn.execute(
            f"UPDATE bid_sov_item SET in_scope = ?, updated_at = CURRENT_TIMESTAMP WHERE id IN ({placeholders})",
            [body.in_scope] + body.item_ids,
        )
        conn.commit()
        return {"updated": len(body.item_ids), "in_scope": body.in_scope}
    finally:
        conn.close()


@router.post("/bids/{bid_id}/sov/set-work-type")
async def set_sov_work_type(bid_id: int, body: SOVWorkTypeUpdate):
    """Set work_type for multiple SOV items at once."""
    valid_types = ("self_perform", "subcontract", "undecided")
    if body.work_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"work_type must be one of {valid_types}")
    conn = get_connection()
    try:
        placeholders = ",".join("?" for _ in body.item_ids)
        conn.execute(
            f"UPDATE bid_sov_item SET work_type = ?, updated_at = CURRENT_TIMESTAMP WHERE id IN ({placeholders})",
            [body.work_type] + body.item_ids,
        )
        conn.commit()
        return {"updated": len(body.item_ids), "work_type": body.work_type}
    finally:
        conn.close()


@router.put("/sov/{item_id}")
async def update_sov_item(item_id: int, item: SOVItemUpdate):
    """Update a SOV item."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_sov_item WHERE id = ?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="SOV item not found")

        updates = {}
        for k, v in item.model_dump().items():
            if v is not None:
                updates[k] = v
        # Allow setting pricing_group_id to None (ungroup)
        if "pricing_group_id" not in updates and item.pricing_group_id is None and "pricing_group_id" in item.model_fields_set:
            updates["pricing_group_id"] = None
        # Allow setting section_id to None (unsection)
        if "section_id" not in updates and item.section_id is None and "section_id" in item.model_fields_set:
            updates["section_id"] = None
        # Allow setting hcss_number to None
        if "hcss_number" not in updates and item.hcss_number is None and "hcss_number" in item.model_fields_set:
            updates["hcss_number"] = None
        # Allow setting holding_description to None
        if "holding_description" not in updates and item.holding_description is None and "holding_description" in item.model_fields_set:
            updates["holding_description"] = None

        if not updates:
            return dict(existing)

        updates["updated_at"] = datetime.now(tz=timezone.utc).isoformat()
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [item_id]
        conn.execute(f"UPDATE bid_sov_item SET {set_clause} WHERE id = ?", values)
        conn.commit()

        row = conn.execute("SELECT * FROM bid_sov_item WHERE id = ?", (item_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/sov/{item_id}")
async def delete_sov_item(item_id: int):
    """Delete a SOV item."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_sov_item WHERE id = ?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="SOV item not found")

        conn.execute("DELETE FROM bid_sov_item WHERE id = ?", (item_id,))
        conn.commit()
        return {"deleted": True, "item_id": item_id}
    finally:
        conn.close()



# ── Pricing Groups ───────────────────────────────────────────────

@router.post("/bids/{bid_id}/groups")
async def create_group(bid_id: int, group: GroupCreate):
    """Create a pricing group."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        cursor = conn.execute(
            "INSERT INTO pricing_group (bid_id, name, description) VALUES (?, ?, ?)",
            (bid_id, group.name, group.description),
        )
        conn.commit()

        row = conn.execute("SELECT * FROM pricing_group WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.get("/bids/{bid_id}/groups")
async def list_groups(bid_id: int):
    """List pricing groups for a bid, with item counts."""
    conn = get_connection()
    try:
        groups = conn.execute(
            "SELECT * FROM pricing_group WHERE bid_id = ? ORDER BY name ASC",
            (bid_id,),
        ).fetchall()

        result = []
        for g in groups:
            gd = dict(g)
            gd["item_count"] = conn.execute(
                "SELECT COUNT(*) as cnt FROM bid_sov_item WHERE pricing_group_id = ?",
                (g["id"],),
            ).fetchone()["cnt"]
            result.append(gd)
        return result
    finally:
        conn.close()


@router.put("/groups/{group_id}")
async def update_group(group_id: int, group: GroupUpdate):
    """Update group name/description."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM pricing_group WHERE id = ?", (group_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Pricing group not found")

        updates = {k: v for k, v in group.model_dump().items() if v is not None}
        if not updates:
            return dict(existing)

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [group_id]
        conn.execute(f"UPDATE pricing_group SET {set_clause} WHERE id = ?", values)
        conn.commit()

        row = conn.execute("SELECT * FROM pricing_group WHERE id = ?", (group_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/groups/{group_id}")
async def delete_group(group_id: int):
    """Delete group and nullify items' pricing_group_id."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM pricing_group WHERE id = ?", (group_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Pricing group not found")

        conn.execute(
            "UPDATE bid_sov_item SET pricing_group_id = NULL WHERE pricing_group_id = ?",
            (group_id,),
        )
        conn.execute("DELETE FROM pricing_group WHERE id = ?", (group_id,))
        conn.commit()
        return {"deleted": True, "group_id": group_id}
    finally:
        conn.close()


@router.post("/groups/{group_id}/assign")
async def assign_items(group_id: int, body: AssignItems):
    """Assign SOV items to a pricing group."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM pricing_group WHERE id = ?", (group_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Pricing group not found")

        for item_id in body.item_ids:
            conn.execute(
                "UPDATE bid_sov_item SET pricing_group_id = ? WHERE id = ?",
                (group_id, item_id),
            )
        conn.commit()
        return {"assigned": len(body.item_ids), "group_id": group_id}
    finally:
        conn.close()


@router.post("/groups/{group_id}/unassign")
async def unassign_items(group_id: int, body: AssignItems):
    """Remove SOV items from a pricing group."""
    conn = get_connection()
    try:
        for item_id in body.item_ids:
            conn.execute(
                "UPDATE bid_sov_item SET pricing_group_id = NULL WHERE id = ? AND pricing_group_id = ?",
                (item_id, group_id),
            )
        conn.commit()
        return {"unassigned": len(body.item_ids), "group_id": group_id}
    finally:
        conn.close()


# ── Sections (replacing pricing groups) ────────────────────────


@router.post("/bids/{bid_id}/sections")
async def create_section(bid_id: int, body: SectionCreate):
    """Create a new section for organizing SOV items."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")

        sort_order = body.sort_order
        if sort_order is None:
            max_sort = conn.execute(
                "SELECT COALESCE(MAX(sort_order), -1) as m FROM bid_section WHERE bid_id = ?",
                (bid_id,),
            ).fetchone()["m"]
            sort_order = max_sort + 1

        cursor = conn.execute(
            "INSERT INTO bid_section (bid_id, name, sort_order) VALUES (?, ?, ?)",
            (bid_id, body.name, sort_order),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM bid_section WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.get("/bids/{bid_id}/sections")
async def list_sections(bid_id: int):
    """List all sections for a bid with item counts."""
    conn = get_connection()
    try:
        sections = conn.execute(
            "SELECT * FROM bid_section WHERE bid_id = ? ORDER BY sort_order ASC",
            (bid_id,),
        ).fetchall()

        result = []
        for s in sections:
            d = dict(s)
            d["item_count"] = conn.execute(
                "SELECT COUNT(*) as cnt FROM bid_sov_item WHERE section_id = ?",
                (s["id"],),
            ).fetchone()["cnt"]
            result.append(d)
        return result
    finally:
        conn.close()


@router.put("/sections/{section_id}")
async def update_section(section_id: int, body: SectionUpdate):
    """Update a section's name, sort_order, or collapsed state."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_section WHERE id = ?", (section_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Section not found")

        updates = {}
        for k, v in body.model_dump().items():
            if v is not None:
                updates[k] = v
        if not updates:
            return dict(existing)

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [section_id]
        conn.execute(f"UPDATE bid_section SET {set_clause} WHERE id = ?", values)
        conn.commit()

        row = conn.execute("SELECT * FROM bid_section WHERE id = ?", (section_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/sections/{section_id}")
async def delete_section(section_id: int):
    """Delete a section, nullifying item section_id refs."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_section WHERE id = ?", (section_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Section not found")

        conn.execute(
            "UPDATE bid_sov_item SET section_id = NULL WHERE section_id = ?",
            (section_id,),
        )
        conn.execute("DELETE FROM bid_section WHERE id = ?", (section_id,))
        conn.commit()
        return {"deleted": True, "section_id": section_id}
    finally:
        conn.close()


@router.post("/sections/{section_id}/assign")
async def assign_items_to_section(section_id: int, body: SectionAssign):
    """Assign SOV items to a section."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM bid_section WHERE id = ?", (section_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Section not found")

        for item_id in body.item_ids:
            conn.execute(
                "UPDATE bid_sov_item SET section_id = ? WHERE id = ?",
                (section_id, item_id),
            )
        conn.commit()
        return {"assigned": len(body.item_ids), "section_id": section_id}
    finally:
        conn.close()


# ── Holding Accounts ──────────────────────────────────────────


@router.post("/bids/{bid_id}/sov/{item_id}/make-holding")
async def make_holding_account(bid_id: int, item_id: int, body: HoldingAccountCreate):
    """Mark an SOV item as a holding account."""
    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT * FROM bid_sov_item WHERE id = ? AND bid_id = ?", (item_id, bid_id)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="SOV item not found")

        conn.execute(
            "UPDATE bid_sov_item SET is_holding_account = 1, holding_description = ? WHERE id = ?",
            (body.holding_description, item_id),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM bid_sov_item WHERE id = ?", (item_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/bids/{bid_id}/sov/{item_id}/make-holding")
async def unmake_holding_account(bid_id: int, item_id: int):
    """Convert a holding account back to a regular item."""
    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT * FROM bid_sov_item WHERE id = ? AND bid_id = ?", (item_id, bid_id)
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="SOV item not found")

        conn.execute(
            "UPDATE bid_sov_item SET is_holding_account = 0, holding_description = NULL WHERE id = ?",
            (item_id,),
        )
        # Remove any distribution mappings
        conn.execute("DELETE FROM holding_distribution WHERE holding_item_id = ?", (item_id,))
        conn.commit()
        row = conn.execute("SELECT * FROM bid_sov_item WHERE id = ?", (item_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.post("/bids/{bid_id}/sov/{item_id}/distribute")
async def set_holding_distribution(bid_id: int, item_id: int, body: HoldingDistribute):
    """Set distribution targets for a holding account."""
    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT * FROM bid_sov_item WHERE id = ? AND bid_id = ? AND is_holding_account = 1",
            (item_id, bid_id),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Holding account not found")

        # Replace all distributions
        conn.execute("DELETE FROM holding_distribution WHERE holding_item_id = ?", (item_id,))
        for target_id in body.target_item_ids:
            conn.execute(
                "INSERT INTO holding_distribution (holding_item_id, target_item_id) VALUES (?, ?)",
                (item_id, target_id),
            )
        conn.commit()
        return {"holding_item_id": item_id, "target_count": len(body.target_item_ids)}
    finally:
        conn.close()


@router.get("/bids/{bid_id}/sov/{item_id}/distribution")
async def get_holding_distribution(bid_id: int, item_id: int):
    """Get current distribution targets for a holding account."""
    conn = get_connection()
    try:
        rows = conn.execute(
            """SELECT hd.target_item_id, s.item_number, s.description
               FROM holding_distribution hd
               JOIN bid_sov_item s ON hd.target_item_id = s.id
               WHERE hd.holding_item_id = ?""",
            (item_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.get("/bids/{bid_id}/holding-accounts")
async def list_holding_accounts(bid_id: int):
    """List all holding accounts with their distributions."""
    conn = get_connection()
    try:
        holdings = conn.execute(
            "SELECT * FROM bid_sov_item WHERE bid_id = ? AND is_holding_account = 1 ORDER BY sort_order",
            (bid_id,),
        ).fetchall()

        result = []
        for h in holdings:
            d = dict(h)
            targets = conn.execute(
                """SELECT hd.target_item_id, s.item_number, s.description
                   FROM holding_distribution hd
                   JOIN bid_sov_item s ON hd.target_item_id = s.id
                   WHERE hd.holding_item_id = ?""",
                (h["id"],),
            ).fetchall()
            d["distribution_targets"] = [dict(t) for t in targets]
            result.append(d)
        return result
    finally:
        conn.close()


# ── Drawing Register ──────────────────────────────────────────


@router.get("/bids/{bid_id}/drawing-register")
async def list_drawing_register(bid_id: int):
    """Returns full drawing register sorted by discipline then drawing_number."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM drawing_register WHERE bid_id = ? ORDER BY discipline ASC, drawing_number ASC",
            (bid_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.put("/drawing-register/{reg_id}")
async def update_drawing_register(reg_id: int, body: DrawingRegisterUpdate):
    """Manual edit of a drawing register entry."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM drawing_register WHERE id = ?", (reg_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Drawing register entry not found")

        updates = {k: v for k, v in body.model_dump().items() if v is not None}
        if not updates:
            return dict(existing)
        updates["ai_generated"] = 0  # manual edit overrides AI
        updates["updated_at"] = datetime.now(tz=timezone.utc).isoformat()

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [reg_id]
        conn.execute(f"UPDATE drawing_register SET {set_clause} WHERE id = ?", values)
        conn.commit()
        row = conn.execute("SELECT * FROM drawing_register WHERE id = ?", (reg_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.post("/bids/{bid_id}/drawing-register/rebuild")
async def rebuild_drawing_register(bid_id: int):
    """Force rebuild drawing register from documents."""
    try:
        from app.services.register_builder import build_drawing_register
        result = build_drawing_register(bid_id)
        return result
    except ImportError:
        return {"status": "ok", "message": "Register builder not available — no API key?", "created": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bids/{bid_id}/drawing-register/export")
async def export_drawing_register(bid_id: int):
    """Export drawing register as Excel."""
    import io
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT drawing_number, title, discipline, revision, addendum_number, date_received, is_new, is_revised, notes FROM drawing_register WHERE bid_id = ? ORDER BY discipline, drawing_number",
            (bid_id,),
        ).fetchall()
    finally:
        conn.close()

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Drawing Register"
        headers = ["Drawing #", "Title", "Discipline", "Rev", "Addendum", "Date Received", "New", "Revised", "Notes"]
        ws.append(headers)
        for r in rows:
            ws.append([r["drawing_number"], r["title"], r["discipline"], r["revision"],
                       r["addendum_number"], r["date_received"],
                       "Yes" if r["is_new"] else "", "Yes" if r["is_revised"] else "", r["notes"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=drawing_register_bid_{bid_id}.xlsx"})
    except ImportError:
        # Fallback to CSV
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Drawing #", "Title", "Discipline", "Rev", "Addendum", "Date Received", "New", "Revised", "Notes"])
        for r in rows:
            writer.writerow([r["drawing_number"], r["title"], r["discipline"], r["revision"],
                             r["addendum_number"], r["date_received"], r["is_new"], r["is_revised"], r["notes"]])
        from fastapi.responses import Response
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=drawing_register_bid_{bid_id}.csv"})


# ── Spec Register ────────────────────────────────────────────


@router.get("/bids/{bid_id}/spec-register")
async def list_spec_register(bid_id: int):
    """Returns full spec register sorted by division then spec_section."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM spec_register WHERE bid_id = ? ORDER BY division ASC, spec_section ASC",
            (bid_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.put("/spec-register/{reg_id}")
async def update_spec_register(reg_id: int, body: SpecRegisterUpdate):
    """Manual edit of a spec register entry."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM spec_register WHERE id = ?", (reg_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Spec register entry not found")

        updates = {k: v for k, v in body.model_dump().items() if v is not None}
        if not updates:
            return dict(existing)
        updates["ai_generated"] = 0
        updates["updated_at"] = datetime.now(tz=timezone.utc).isoformat()

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [reg_id]
        conn.execute(f"UPDATE spec_register SET {set_clause} WHERE id = ?", values)
        conn.commit()
        row = conn.execute("SELECT * FROM spec_register WHERE id = ?", (reg_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.post("/bids/{bid_id}/spec-register/rebuild")
async def rebuild_spec_register(bid_id: int):
    """Force rebuild spec register from documents."""
    try:
        from app.services.register_builder import build_spec_register
        result = build_spec_register(bid_id)
        return result
    except ImportError:
        return {"status": "ok", "message": "Register builder not available", "created": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bids/{bid_id}/spec-register/export")
async def export_spec_register(bid_id: int):
    """Export spec register as Excel."""
    import io
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT spec_section, title, division, addendum_number, date_received, is_new, is_revised, revision_summary, notes FROM spec_register WHERE bid_id = ? ORDER BY division, spec_section",
            (bid_id,),
        ).fetchall()
    finally:
        conn.close()

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spec Register"
        headers = ["Section #", "Title", "Division", "Addendum", "Date Received", "New", "Revised", "Revision Summary", "Notes"]
        ws.append(headers)
        for r in rows:
            ws.append([r["spec_section"], r["title"], r["division"], r["addendum_number"],
                       r["date_received"], "Yes" if r["is_new"] else "", "Yes" if r["is_revised"] else "",
                       r["revision_summary"], r["notes"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=spec_register_bid_{bid_id}.xlsx"})
    except ImportError:
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Section #", "Title", "Division", "Addendum", "Date Received", "New", "Revised", "Revision Summary", "Notes"])
        for r in rows:
            writer.writerow([r["spec_section"], r["title"], r["division"], r["addendum_number"],
                             r["date_received"], r["is_new"], r["is_revised"], r["revision_summary"], r["notes"]])
        from fastapi.responses import Response
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=spec_register_bid_{bid_id}.csv"})


# ── RFI & Clarifications Log ─────────────────────────────────


@router.get("/bids/{bid_id}/rfi-log")
async def list_rfi_log(bid_id: int):
    """Returns full RFI log sorted by addendum then rfi_number."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM rfi_log WHERE bid_id = ? ORDER BY addendum_number ASC, rfi_number ASC",
            (bid_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.post("/bids/{bid_id}/rfi-log")
async def add_rfi_entry(bid_id: int, body: RFICreate):
    """Manually add an RFI entry."""
    conn = get_connection()
    try:
        cursor = conn.execute(
            """INSERT INTO rfi_log (bid_id, rfi_number, question, asked_by, date_asked,
               response, responded_by, date_responded, addendum_number, related_spec,
               related_drawing, status, notes, ai_generated)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)""",
            (bid_id, body.rfi_number, body.question, body.asked_by, body.date_asked,
             body.response, body.responded_by, body.date_responded, body.addendum_number,
             body.related_spec, body.related_drawing, body.status or "answered", body.notes),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM rfi_log WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.put("/rfi-log/{rfi_id}")
async def update_rfi_entry(rfi_id: int, body: RFIUpdate):
    """Edit an RFI entry."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM rfi_log WHERE id = ?", (rfi_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="RFI entry not found")

        updates = {k: v for k, v in body.model_dump().items() if v is not None}
        if not updates:
            return dict(existing)
        updates["updated_at"] = datetime.now(tz=timezone.utc).isoformat()

        set_clause = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [rfi_id]
        conn.execute(f"UPDATE rfi_log SET {set_clause} WHERE id = ?", values)
        conn.commit()
        row = conn.execute("SELECT * FROM rfi_log WHERE id = ?", (rfi_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@router.delete("/rfi-log/{rfi_id}")
async def delete_rfi_entry(rfi_id: int):
    """Delete an RFI entry."""
    conn = get_connection()
    try:
        existing = conn.execute("SELECT * FROM rfi_log WHERE id = ?", (rfi_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="RFI entry not found")
        conn.execute("DELETE FROM rfi_log WHERE id = ?", (rfi_id,))
        conn.commit()
        return {"deleted": True, "rfi_id": rfi_id}
    finally:
        conn.close()


@router.post("/bids/{bid_id}/rfi-log/rebuild")
async def rebuild_rfi_log(bid_id: int):
    """Force rebuild RFI log from documents."""
    try:
        from app.services.rfi_parser import build_rfi_log
        result = build_rfi_log(bid_id)
        return result
    except ImportError:
        return {"status": "ok", "message": "RFI parser not available", "created": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bids/{bid_id}/rfi-log/export")
async def export_rfi_log(bid_id: int):
    """Export RFI log as Excel."""
    import io
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT rfi_number, question, asked_by, date_asked, response, responded_by, date_responded, addendum_number, related_spec, related_drawing, status, notes FROM rfi_log WHERE bid_id = ? ORDER BY addendum_number, rfi_number",
            (bid_id,),
        ).fetchall()
    finally:
        conn.close()

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFI Log"
        headers = ["RFI #", "Question", "Asked By", "Date Asked", "Response", "Responded By", "Date Responded", "Addendum", "Related Spec", "Related Drawing", "Status", "Notes"]
        ws.append(headers)
        for r in rows:
            ws.append([r["rfi_number"], r["question"], r["asked_by"], r["date_asked"],
                       r["response"], r["responded_by"], r["date_responded"], r["addendum_number"],
                       r["related_spec"], r["related_drawing"], r["status"], r["notes"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=rfi_log_bid_{bid_id}.xlsx"})
    except ImportError:
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["RFI #", "Question", "Asked By", "Date Asked", "Response", "Responded By", "Date Responded", "Addendum", "Related Spec", "Related Drawing", "Status", "Notes"])
        for r in rows:
            writer.writerow([r["rfi_number"], r["question"], r["asked_by"], r["date_asked"],
                             r["response"], r["responded_by"], r["date_responded"], r["addendum_number"],
                             r["related_spec"], r["related_drawing"], r["status"], r["notes"]])
        from fastapi.responses import Response
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=rfi_log_bid_{bid_id}.csv"})


# ── Document Chunking ──────────────────────────────────────────

@router.post("/bids/{bid_id}/rechunk")
async def rechunk_documents(bid_id: int):
    """Delete existing chunks and re-chunk all documents for a bid."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")
    finally:
        conn.close()

    result = chunk_all_bid_documents(bid_id)

    # Rebuild vector index for this bid after re-chunking
    if VECTOR_SEARCH_ENABLED:
        try:
            from app.services.vector_store import rebuild_bid_index
            rebuild_bid_index(bid_id)
        except Exception as e:
            logger.warning("Vector index rebuild failed for bid %d: %s", bid_id, e)

    return result


# ── Historical Rate Lookup ─────────────────────────────────────

class RateLookupRequest(BaseModel):
    description: str
    unit: Optional[str] = None
    quantity: Optional[float] = None


@router.post("/bids/{bid_id}/sov/{item_id}/lookup")
async def lookup_sov_rates(bid_id: int, item_id: int):
    """Find historical rates for a specific SOV item."""
    conn = get_connection()
    try:
        item = conn.execute(
            "SELECT * FROM bid_sov_item WHERE id = ? AND bid_id = ?",
            (item_id, bid_id),
        ).fetchone()
        if not item:
            raise HTTPException(status_code=404, detail="SOV item not found")
    finally:
        conn.close()

    from dataclasses import asdict
    matches = lookup_rates_for_sov_item(
        item["description"], item["unit"], item["quantity"],
    )
    return {"item_id": item_id, "matches": [asdict(m) for m in matches]}


@router.post("/bids/{bid_id}/sov/auto-rate")
async def auto_rate_sov(bid_id: int):
    """Auto-populate historical rates for all SOV items in a bid."""
    conn = get_connection()
    try:
        bid = conn.execute("SELECT id FROM active_bids WHERE id = ?", (bid_id,)).fetchone()
        if not bid:
            raise HTTPException(status_code=404, detail="Bid not found")
    finally:
        conn.close()

    result = auto_populate_sov_rates(bid_id)
    return result


@router.get("/bids/{bid_id}/sov/{item_id}/rates")
async def get_sov_rates(bid_id: int, item_id: int):
    """Get rate match candidates for a SOV item (cached from last lookup)."""
    conn = get_connection()
    try:
        item = conn.execute(
            "SELECT id, description, unit, quantity, rate_source, rate_confidence, unit_price "
            "FROM bid_sov_item WHERE id = ? AND bid_id = ?",
            (item_id, bid_id),
        ).fetchone()
        if not item:
            raise HTTPException(status_code=404, detail="SOV item not found")
        return dict(item)
    finally:
        conn.close()
